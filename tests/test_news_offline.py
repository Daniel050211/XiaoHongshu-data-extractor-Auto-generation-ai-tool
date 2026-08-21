"""佛山新聞 AI 離線測試：不連網、不呼叫真實 API（NEWS_AI_FAKE=1）。"""
from __future__ import annotations

import os
import tempfile
import time
import urllib.parse
import urllib.request
import unittest
from dataclasses import replace
from pathlib import Path
from unittest import mock

os.environ["NEWS_AI_FAKE"] = "1"

from news_app import account_store, ai, email as news_email, mailwatch, pipeline, prompts, store, web  # noqa: E402
from news_app.config import NewsAccount, NewsConfig, _as_list  # noqa: E402

PROJECT_ROOT = Path(__file__).resolve().parent.parent
FIXTURE = PROJECT_ROOT / "data" / "fixtures" / "news_serper_sample.json"


class NewsPipelineTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmp = Path(tempfile.mkdtemp(prefix="news_test_"))
        base = NewsConfig.load()
        cls.cfg = replace(base, db_path=cls.tmp / "news.db", data_dir=cls.tmp)

    def setUp(self):
        self.conn = store.connect(self.cfg.db_path)
        # 每筆測試獨立 run；共用資料庫沒關係，用 run_id 區分

    def tearDown(self):
        self.conn.close()

    def _start(self):
        run_id = pipeline.start_run(
            self.cfg, self.conn,
            run_date="2026-08-17",
            from_json=FIXTURE,
            dry_run=True,
            notify=False,
        )
        return run_id

    def test_full_flow_approve(self):
        run_id = self._start()
        run = store.get_run(self.conn, run_id)
        self.assertEqual(run["status"], pipeline.STATUS_AWAIT_DIRECTION)
        self.assertEqual(run["articles_count"], 6)
        directions = store.get_directions(self.conn, run_id)
        self.assertEqual(len(directions), 3)
        self.assertTrue(directions[0]["title"])
        self.assertTrue(directions[0]["sources"])

        status = pipeline.decide_direction(self.cfg, self.conn, run_id, "2", dry_run=True, notify=False)
        self.assertEqual(status, pipeline.STATUS_AWAIT_SCRIPT)
        run = store.get_run(self.conn, run_id)
        self.assertIn("核心觀點", run["analysis"])
        versions = store.get_versions(self.conn, run_id)
        self.assertEqual(len(versions), 3)
        self.assertEqual([v["style"] for v in versions], ["反差型", "數據型", "判斷型"])

        status = pipeline.decide_script(self.cfg, self.conn, run_id, "3", dry_run=True, notify=False)
        self.assertEqual(status, pipeline.STATUS_DONE)
        run = store.get_run(self.conn, run_id)
        self.assertTrue(run["tagline"])
        self.assertTrue(run["image_prompt"])
        self.assertEqual(run["style"], "判斷型")
        self.assertIn("佛山", run["script_to_publish"])

    def test_direction_reject_then_retry(self):
        run_id = self._start()
        status = pipeline.decide_direction(
            self.cfg, self.conn, run_id, "reject", "方向太籠統", dry_run=True, notify=False)
        self.assertEqual(status, pipeline.STATUS_AWAIT_DIRECTION)
        run = store.get_run(self.conn, run_id)
        self.assertEqual(run["retry_direction"], 1)
        # 拒絕後重新產生了新方向
        self.assertEqual(len(store.get_directions(self.conn, run_id)), 3)

        status = pipeline.decide_direction(self.cfg, self.conn, run_id, "1", dry_run=True, notify=False)
        self.assertEqual(status, pipeline.STATUS_AWAIT_SCRIPT)

    def test_script_reject_then_approve(self):
        run_id = self._start()
        pipeline.decide_direction(self.cfg, self.conn, run_id, "1", dry_run=True, notify=False)
        status = pipeline.decide_script(
            self.cfg, self.conn, run_id, "reject", "語氣再克制一點", dry_run=True, notify=False)
        self.assertEqual(status, pipeline.STATUS_AWAIT_SCRIPT)
        run = store.get_run(self.conn, run_id)
        self.assertEqual(run["retry_script"], 1)
        status = pipeline.decide_script(self.cfg, self.conn, run_id, "反差型", dry_run=True, notify=False)
        self.assertEqual(status, pipeline.STATUS_DONE)

    def test_max_direction_retries_fails(self):
        cfg = replace(self.cfg, direction_max_retries=1)
        run_id = self._start()
        pipeline.decide_direction(cfg, self.conn, run_id, "reject", "意見1", dry_run=True, notify=False)
        status = pipeline.decide_direction(cfg, self.conn, run_id, "reject", "意見2", dry_run=True, notify=False)
        self.assertEqual(status, pipeline.STATUS_FAILED)
        run = store.get_run(self.conn, run_id)
        self.assertIn("最大重試次數", run["error"])

    def test_wrong_state_rejected(self):
        run_id = self._start()
        with self.assertRaises(ValueError):
            pipeline.decide_script(self.cfg, self.conn, run_id, "1", dry_run=True, notify=False)


class AiParseTest(unittest.TestCase):
    def test_parse_directions_messy(self):
        text = """```json
{
  "news_summary": "摘要",
  "directions": [
    {"id": "d1", "title": "標題", "description": "說明", "sources": [{"title": "s", "url": "u"}]},
    {"id": "d2", "title": "標題2", "description": "說明2", "sources": [],}
  ],
}
```"""
        directions, summary = ai.parse_directions(text)
        self.assertEqual(summary, "摘要")
        self.assertEqual(len(directions), 2)

    def test_parse_tagline_fallback(self):
        tagline, prompt = ai.parse_tagline(
            '{"tagline": "一句話", "image_prompt": "圖片描述"}')
        self.assertEqual(tagline, "一句話")
        self.assertTrue(prompt)

    def test_chat_json_retries_on_bad_output(self):
        bad = '{"versions": [{"style": "反差型", "content": "內容有"未跳脫引號"且未閉合'
        good = '{"versions": [{"style": "反差型", "content": "完整內容"}]}'
        calls = {"n": 0}

        def fake_chat(cfg, system, user, **kw):
            calls["n"] += 1
            return good if calls["n"] > 1 else bad

        with mock.patch("news_app.ai.chat", side_effect=fake_chat):
            versions = ai.chat_json(None, "sys", "user", ai.parse_versions, attempts=2)
        self.assertEqual(calls["n"], 2)
        self.assertEqual(versions[0]["style"], "反差型")

    def test_chat_json_gives_up_after_attempts(self):
        with mock.patch("news_app.ai.chat", return_value="不是 JSON"):
            with self.assertRaises(ValueError):
                ai.chat_json(None, "sys", "user", ai.parse_versions, attempts=2)

    def test_retry_scripts_recovers(self):
        tmp = Path(tempfile.mkdtemp(prefix="news_retry_"))
        cfg = replace(NewsConfig.load(), db_path=tmp / "news.db")
        conn = store.connect(cfg.db_path)
        run_id = pipeline.start_run(cfg, conn, run_date="2026-08-17",
                                    from_json=FIXTURE, dry_run=True, notify=False)
        pipeline.decide_direction(cfg, conn, run_id, "1", dry_run=True, notify=False)
        run = store.get_run(conn, run_id)
        self.assertEqual(run["status"], pipeline.STATUS_AWAIT_SCRIPT)
        # 模擬腳本階段失敗後，用 retry_scripts 復原
        store.set_run_status(conn, run_id, pipeline.STATUS_FAILED, "模擬失敗")
        status = pipeline.retry_scripts(cfg, conn, run_id, dry_run=True, notify=False)
        self.assertEqual(status, pipeline.STATUS_AWAIT_SCRIPT)
        self.assertEqual(len(store.get_versions(conn, run_id)), 3)
        conn.close()


class MultiAccountTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmp = Path(tempfile.mkdtemp(prefix="news_multi_"))
        base = NewsConfig.load()
        cls.cfg = replace(
            base,
            db_path=cls.tmp / "news.db",
            accounts=[
                NewsAccount(name="佛山產業號", query="產業 新聞", audience="打工人",
                            topics="佛山AI、機器人"),
                NewsAccount(name="旅遊號", query="佛山 旅遊 美食", audience="旅行愛好者",
                            topics="佛山旅遊、美食", email_to=["travel@example.com"]),
            ],
        )

    def setUp(self):
        self.conn = store.connect(self.cfg.db_path)

    def tearDown(self):
        self.conn.close()

    def test_prompt_per_account(self):
        acc = self.cfg.enabled_accounts()[1]
        sys_text = prompts.select_directions_system(acc)
        self.assertIn("目標讀者是旅行愛好者", sys_text)
        self.assertIn("佛山旅遊、美食", sys_text)
        self.assertIn("旅行愛好者", prompts.script_system(acc))
        self.assertIn("佛山旅遊、美食", prompts.tagline_system(acc))

    def test_custom_prompt_override(self):
        acc = NewsAccount(name="x", audience="a", topics="b",
                          prompt_directions="完全自訂方向 prompt")
        self.assertEqual(prompts.select_directions_system(acc), "完全自訂方向 prompt")
        # 沒填 prompt_analysis → 仍套用 audience/topics 模板替換
        self.assertIn("目標讀者是a", prompts.deep_analysis_system(acc))
        acc2 = NewsAccount(name="y", prompt_scripts="自訂腳本 prompt")
        self.assertEqual(prompts.script_system(acc2), "自訂腳本 prompt")
        acc3 = NewsAccount(name="z", prompt_tagline="自訂 tagline prompt")
        self.assertEqual(prompts.tagline_system(acc3), "自訂 tagline prompt")

    def test_non_foshan_account_prompt_clean(self):
        acc = NewsAccount(name="台北美食號", place="台北", audience="愛吃宵夜的年輕人",
                          topics="台北夜市、小吃", tone="熱情、有畫面",
                          hashtags="#台北美食")
        for text in (
            prompts.select_directions_system(acc),
            prompts.deep_analysis_system(acc),
            prompts.script_system(acc),
            prompts.tagline_system(acc),
        ):
            self.assertIn("台北", text)
            self.assertNotIn("佛山", text)
        self.assertIn("對愛吃宵夜的年輕人的啟示", prompts.deep_analysis_system(acc))
        self.assertNotIn("對打工人的啟示", prompts.deep_analysis_system(acc))

    def test_run_account_stored_and_uses_own_query(self):
        run_id = pipeline.start_run(
            self.cfg, self.conn, run_date="2026-08-17",
            from_json=FIXTURE, dry_run=True, notify=False, account="旅遊號")
        run = store.get_run(self.conn, run_id)
        self.assertEqual(run["account"], "旅遊號")
        self.assertEqual(run["query"], "佛山 旅遊 美食")

    def test_disabled_account_resolvable_by_name(self):
        cfg = replace(
            self.cfg,
            accounts=[
                NewsAccount(name="產業號", query="產業"),
                NewsAccount(name="旅遊號", enabled=False, query="旅遊 美食"),
            ],
        )
        run_id = pipeline.start_run(
            cfg, self.conn, run_date="2026-08-17",
            from_json=FIXTURE, dry_run=True, notify=False, account="旅遊號")
        run = store.get_run(self.conn, run_id)
        self.assertEqual(run["account"], "旅遊號")
        self.assertEqual(run["query"], "旅遊 美食")

    def test_accounts_dir_isolated_and_bad_file_skipped(self):
        tmp = Path(tempfile.mkdtemp(prefix="news_accdir_"))
        (tmp / "旅遊號.yaml").write_text(
            "name: 旅遊號\nquery: 旅遊 美食\naudience: 年輕人\n", encoding="utf-8")
        (tmp / "壞檔.yaml").write_text("name: [broken\n", encoding="utf-8")
        accounts = NewsConfig._load_accounts_from_dir(tmp)
        names = [a.name for a in accounts]
        self.assertIn("旅遊號", names)
        self.assertNotIn("壞檔", names)
        # 修改旅遊號的設定，其他帳號不受影響
        (tmp / "旅遊號.yaml").write_text(
            "name: 旅遊號\nquery: 旅遊 美食\naudience: 親子家庭\n", encoding="utf-8")
        accounts2 = NewsConfig._load_accounts_from_dir(tmp)
        travel = [a for a in accounts2 if a.name == "旅遊號"][0]
        self.assertEqual(travel.audience, "親子家庭")

    def test_memory_saved_per_account(self):
        run_id = pipeline.start_run(
            self.cfg, self.conn, run_date="2026-08-17",
            from_json=FIXTURE, dry_run=True, notify=False, account="旅遊號")
        pipeline.decide_direction(self.cfg, self.conn, run_id, "1", dry_run=True, notify=False)
        pipeline.decide_script(self.cfg, self.conn, run_id, "2", dry_run=True, notify=False)
        memory = store.latest_memory(self.conn, "旅遊號")
        self.assertIn("Tagline", memory)
        self.assertEqual(store.latest_memory(self.conn, "佛山產業號"), "")


class WebFormTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        tmp = Path(tempfile.mkdtemp(prefix="news_web_"))
        base = NewsConfig.load()
        # 明確清空公開網址/安全碼，避免 .env 的 FORM_PUBLIC_URL 把測試導到公開隧道
        cls.cfg = replace(base, db_path=tmp / "news.db", web_port=18999,
                          form_public_url="", form_token="", data_dir=tmp)
        cls.conn = store.connect(cls.cfg.db_path)
        cls.run_id = pipeline.start_run(
            cls.cfg, cls.conn, run_date="2026-08-17",
            from_json=FIXTURE, dry_run=True, notify=False)
        cls.httpd, cls.thread = web.start_server(cls.cfg)

    @classmethod
    def tearDownClass(cls):
        cls.httpd.shutdown()
        cls.thread.join(timeout=5)
        cls.conn.close()

    def test_form_get_and_post(self):
        url = web.approval_url(self.cfg, self.run_id)
        with urllib.request.urlopen(url, timeout=10) as r:
            page = r.read().decode("utf-8")
        self.assertIn("方向 1", page)
        self.assertIn("送出審批", page)
        self.assertIn("拒絕全部", page)

        real_decide = pipeline.decide_direction

        def offline_decide(cfg, conn, run_id, decision, comment="", dry_run=False, notify=True):
            return real_decide(cfg, conn, run_id, decision, comment, dry_run=True, notify=False)

        data = urllib.parse.urlencode({"decision": "1", "comment": ""}).encode()
        patcher = mock.patch("news_app.web.pipeline.decide_direction", side_effect=offline_decide)
        patcher.start()
        try:
            req = urllib.request.Request(url, data=data)
            with urllib.request.urlopen(req, timeout=10) as r:
                r.read()

            for _ in range(30):
                run = store.get_run(self.conn, self.run_id)
                if run["status"] in (pipeline.STATUS_AWAIT_SCRIPT, pipeline.STATUS_DONE,
                                     pipeline.STATUS_FAILED):
                    break
                time.sleep(1)
        finally:
            patcher.stop()
        self.assertEqual(run["status"], pipeline.STATUS_AWAIT_SCRIPT)

        with urllib.request.urlopen(url, timeout=10) as r:
            page2 = r.read().decode("utf-8")
        self.assertIn("反差型", page2)
        self.assertIn("數據型", page2)


class WebFormTokenTest(unittest.TestCase):
    """手機公開表單：安全碼驗證與公開網址。"""

    @classmethod
    def setUpClass(cls):
        tmp = Path(tempfile.mkdtemp(prefix="news_web_token_"))
        base = NewsConfig.load()
        cls.cfg = replace(
            base,
            db_path=tmp / "news.db",
            web_port=18998,
            data_dir=tmp,
            form_public_url="https://example.ngrok.app",
            form_token="secret123",
        )
        cls.conn = store.connect(cls.cfg.db_path)
        cls.run_id = pipeline.start_run(
            cls.cfg, cls.conn, run_date="2026-08-17",
            from_json=FIXTURE, dry_run=True, notify=False)
        cls.httpd, cls.thread = web.start_server(cls.cfg)

    @classmethod
    def tearDownClass(cls):
        cls.httpd.shutdown()
        cls.thread.join(timeout=5)
        cls.conn.close()

    def test_public_url_has_token(self):
        url = web.approval_url(self.cfg, self.run_id)
        self.assertTrue(url.startswith("https://example.ngrok.app/approve/"))
        self.assertIn("?token=secret123", url)

    def test_without_token_is_403(self):
        local = f"http://127.0.0.1:{self.cfg.web_port}/approve/{self.run_id}"
        with self.assertRaises(urllib.error.HTTPError) as ctx:
            urllib.request.urlopen(local, timeout=10)
        self.assertEqual(ctx.exception.code, 403)

    def test_wrong_token_is_403(self):
        local = f"http://127.0.0.1:{self.cfg.web_port}/approve/{self.run_id}?token=wrong"
        with self.assertRaises(urllib.error.HTTPError) as ctx:
            urllib.request.urlopen(local, timeout=10)
        self.assertEqual(ctx.exception.code, 403)

    def test_correct_token_opens_form(self):
        url = f"http://127.0.0.1:{self.cfg.web_port}/approve/{self.run_id}?token=secret123"
        with urllib.request.urlopen(url, timeout=10) as r:
            page = r.read().decode("utf-8")
        self.assertIn("方向 1", page)
        self.assertIn("送出審批", page)


class AccountStoreTest(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="news_accstore_"))

    def test_roundtrip_toggle_delete(self):
        account_store.save_account({
            "name": "測試號", "enabled": True, "place": "香港",
            "query": "測試 新聞", "audience": "學生", "topics": "學習",
            "email_to": ["a@x.com", "b@x.com"],
        }, path=self.tmp)
        items = account_store.list_accounts(self.tmp)
        self.assertEqual(len(items), 1)
        self.assertEqual(items[0]["name"], "測試號")
        self.assertEqual(items[0]["email_to"], ["a@x.com", "b@x.com"])

        account_store.toggle_enabled("測試號", False, path=self.tmp)
        self.assertFalse(account_store.list_accounts(self.tmp)[0]["enabled"])

        account_store.delete_account("測試號", path=self.tmp)
        self.assertEqual(account_store.list_accounts(self.tmp), [])

    def test_invalid_name_rejected(self):
        with self.assertRaises(ValueError):
            account_store.save_account({"name": "a/b"}, path=self.tmp)
        with self.assertRaises(ValueError):
            account_store.validate_name("   ")

    def test_as_list_splits_string_not_characters(self):
        self.assertEqual(_as_list("a@x.com, b@y.com"), ["a@x.com", "b@y.com"])
        self.assertEqual(_as_list("a@x.com;b@y.com"), ["a@x.com", "b@y.com"])
        self.assertEqual(_as_list(""), [])


class AccountLinkingTest(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="news_link_"))
        base = NewsConfig.load()
        self.cfg = replace(base, db_path=self.tmp / "news.db", data_dir=self.tmp,
                           accounts=[NewsAccount(name="新聞號", xhs_account="XHS源頭")])
        self.conn = store.connect(self.cfg.db_path)

    def tearDown(self):
        self.conn.close()

    def test_news_feedback_uses_xhs_account_binding(self):
        run_id = pipeline.start_run(self.cfg, self.conn, run_date="2026-08-17",
                                    from_json=FIXTURE, dry_run=True, notify=False, account="新聞號")
        captured = {}

        def fake_feedback(account=""):
            captured["account"] = account
            return "回饋內容"

        with mock.patch("news_app.store.latest_feedback_from_xhs", side_effect=fake_feedback):
            pipeline.decide_direction(self.cfg, self.conn, run_id, "1",
                                      dry_run=True, notify=False)
        self.assertEqual(captured.get("account"), "XHS源頭")
        events = store.run_events(self.conn, run_id)
        self.assertTrue(any("XHS源頭" in e["message"] for e in events))

    def test_cancel_run(self):
        run_id = pipeline.start_run(self.cfg, self.conn, run_date="2026-08-17",
                                    from_json=FIXTURE, dry_run=True, notify=False, account="新聞號")
        self.assertTrue(pipeline.cancel_run(self.conn, run_id))
        run = store.get_run(self.conn, run_id)
        self.assertEqual(run["status"], "cancelled")
        self.assertFalse(pipeline.cancel_run(self.conn, run_id))  # 已是終止狀態

    def test_xhs_account_email_to_preserved(self):
        from xhs_report.config import Config as XhsConfig
        cfg_path = self.tmp / "config.yaml"
        cfg_path.write_text(
            "accounts:\n"
            "  - name: A\n"
            "    excel_path: x.xlsx\n"
            "    email_to: [a@b.c, d@e.f]\n",
            encoding="utf-8",
        )
        xc = XhsConfig.load(cfg_path)
        self.assertEqual(xc.accounts[0]["email_to"], ["a@b.c", "d@e.f"])


class MailWatchTest(unittest.TestCase):
    def test_parse_direction_reply(self):
        parsed = mailwatch.parse_reply(
            "AI分析方向選擇 - 2026-08-17（#3）",
            "方向2\n\n第二行開始寫意見\n-----Original Message-----",
        )
        self.assertEqual(parsed["run_id"], 3)
        self.assertEqual(parsed["decision"], "2")
        self.assertEqual(parsed["comment"], "第二行開始寫意見")

    def test_parse_reject_reply(self):
        parsed = mailwatch.parse_reply("內容審核 - 2026-08-17 10:00（#9）", "拒絕全部\n太籠統了")
        self.assertEqual(parsed["run_id"], 9)
        self.assertEqual(parsed["decision"], "reject")
        self.assertEqual(parsed["comment"], "太籠統了")

    def test_parse_script_style_reply(self):
        self.assertEqual(mailwatch.parse_reply("內容審核（#4）", "版本3")["decision"], "3")
        self.assertEqual(mailwatch.parse_reply("內容審核（#4）", "反差型")["decision"], "1")
        self.assertEqual(mailwatch.parse_reply("內容審核（#4）", "1、")["decision"], "1")

    def test_original_approval_email_not_treated_as_reply(self):
        # 自己寄出的審批信（內文是完整說明）不該被誤判成回覆
        body = "📱 小紅書帖子草稿（3個版本）\n請審核並選擇想要的版本…\n1. 反差型\n2. 數據型\n3. 判斷型"
        self.assertIsNone(mailwatch.parse_reply("內容審核 - 2026-08-18 12:27（#14）", body))

    def test_is_self_sent_approval(self):
        run = {"direction_email": "<h2>📊 深度分析方向選擇</h2>\n方向1", "script_email": ""}
        self.assertTrue(mailwatch.is_self_sent_approval("<h2>📊 深度分析方向選擇</h2>\n方向1", run))
        self.assertFalse(mailwatch.is_self_sent_approval("方向1\n請做深度分析", run))

    def test_clean_body_removes_quote_and_signature(self):
        body = ("方向1\n\n補充意見\n\n"
                "> 原始內容\n"
                "--\n簽名檔")
        self.assertNotIn("原始內容", mailwatch.clean_body(body))
        self.assertNotIn("簽名檔", mailwatch.clean_body(body))

    def test_email_contains_reply_instructions(self):
        cfg = replace(NewsConfig.load(), db_path=Path(tempfile.mkdtemp(prefix="news_mail_")) / "news.db")
        conn = store.connect(cfg.db_path)
        run_id = pipeline.start_run(cfg, conn, run_date="2026-08-17",
                                    from_json=FIXTURE, dry_run=True, notify=False)
        run = store.get_run(conn, run_id)
        self.assertIn("直接回覆 Email", run["direction_email"])
        self.assertIn("方向1", run["direction_email"])
        conn.close()

    def test_email_send_accepts_recipients(self):
        cfg = NewsConfig.load()
        cfg.email_to = ["default@x.com"]
        with mock.patch("news_app.email.emailer.send", return_value=True) as m:
            ok = news_email.send(cfg, "測試主旨", "<p>body</p>", recipients=["a@b.c"])
        self.assertTrue(ok)
        args, _kwargs = m.call_args
        self.assertEqual(args[0].email_to, ["a@b.c"])


class SchedulerHelperTest(unittest.TestCase):
    def test_parse_days(self):
        from news_app import scheduler
        self.assertEqual(scheduler.parse_days("mon,wed"), ["Monday", "Wednesday"])
        self.assertEqual(scheduler.parse_days("Monday;Friday"), ["Monday", "Friday"])
        self.assertEqual(scheduler.parse_days(""), [])
        self.assertEqual(scheduler.parse_days("一、三"), ["Monday", "Wednesday"])
        self.assertEqual(scheduler.parse_days("周一,周五"), ["Monday", "Friday"])

    def test_normalize_and_display_days(self):
        from news_app import scheduler
        self.assertEqual(scheduler.normalize_days("一,三"), "mon,wed")
        self.assertEqual(scheduler.normalize_days("週一；週三"), "mon,wed")
        self.assertEqual(scheduler.display_days("mon,wed"), "一、三")
        self.assertEqual(scheduler.normalize_days(""), "")

    def test_default_accounts_skip_scheduled(self):
        from news_app import scheduler
        accounts = [
            {"name": "A", "enabled": True, "schedule_time": "11:00", "schedule_days": "wed"},
            {"name": "B", "enabled": True, "schedule_time": ""},
            {"name": "C", "enabled": False, "schedule_time": "14:00"},
        ]
        self.assertEqual([a["name"] for a in scheduler.scheduled_accounts(accounts)], ["A"])
        self.assertEqual([a["name"] for a in scheduler.default_accounts(accounts)], ["B"])


class ExcelExportTest(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="news_excel_"))
        base = NewsConfig.load()
        self.cfg = replace(base, db_path=self.tmp / "news.db", data_dir=self.tmp,
                           excel_export=True)

    def test_articles_and_final_export(self):
        from openpyxl import load_workbook
        from news_app import excel_export

        articles = [
            {"id": "a1", "topic": "AI", "title": "標題一", "url": "https://x.com/1",
             "snippet": "摘要一", "source": "來源A", "date": "2026-08-20"},
            {"id": "a2", "topic": "機器人", "title": "標題二", "url": "https://x.com/2",
             "snippet": "摘要二", "source": "來源B", "date": "2026-08-21"},
        ]
        excel_export.save_articles(self.cfg, articles, run_id=7, run_date="2026-08-21",
                                   account="佛山科创观察")
        excel_export.save_final(
            self.cfg, run_id=7, run_date="2026-08-21", account="佛山科创观察",
            chosen_direction='{"title": "方向一", "description": "說明"}',
            style="判斷型", script="定稿內容", tagline="一句話", image_prompt="圖片提示",
        )

        path = excel_export.export_path(self.cfg)
        self.assertTrue(path.exists())
        wb = load_workbook(path)
        self.assertIn("新聞", wb.sheetnames)
        self.assertIn("定稿", wb.sheetnames)
        art_rows = list(wb["新聞"].iter_rows(values_only=True))
        self.assertEqual(len(art_rows), 3)  # header + 2 articles
        self.assertEqual(art_rows[1][4], "AI")       # 主題分類
        self.assertEqual(art_rows[2][5], "標題二")    # 標題
        fin_rows = list(wb["定稿"].iter_rows(values_only=True))
        self.assertEqual(len(fin_rows), 2)           # header + 1 final
        self.assertEqual(fin_rows[1][3], "方向一")
        self.assertEqual(fin_rows[1][5], "判斷型")
        self.assertEqual(fin_rows[1][6], "定稿內容")
        self.assertEqual(fin_rows[1][8], "圖片提示")

    def test_disabled_export_writes_nothing(self):
        from news_app import excel_export
        self.cfg = replace(self.cfg, excel_export=False)
        excel_export.save_articles(self.cfg, [], run_id=1, run_date="2026-08-21", account="A")
        self.assertFalse(excel_export.export_path(self.cfg).exists())


if __name__ == "__main__":
    unittest.main()
