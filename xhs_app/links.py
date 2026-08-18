"""讀寫帖文連結 Excel。"""
from __future__ import annotations

from pathlib import Path

import openpyxl


def read_urls(path: str | Path) -> list[str]:
    p = Path(path)
    if not p.exists():
        return []
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb.active
    urls = []
    for row in ws.iter_rows(values_only=True):
        v = row[0] if row else None
        if v and str(v).strip().startswith("http"):
            urls.append(str(v).strip())
    wb.close()
    return urls


def save_urls(path: str | Path, urls: list[str]) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "posts"
    ws.append(["post_url", "備註"])
    for u in urls:
        ws.append([u, ""])
    wb.save(p)
