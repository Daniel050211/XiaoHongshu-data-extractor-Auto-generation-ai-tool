"""把搜到的新聞與最終定稿（方向 + 腳本）自動存成 Excel，方便回顧與報表。"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

ARTICLES_SHEET = "新聞"
FINAL_SHEET = "定稿"
ARTICLES_HEADERS = ["日期", "帳號", "Run#", "序號", "主題分類", "標題", "連結", "摘要", "來源", "發布日期"]
FINAL_HEADERS = ["日期", "帳號", "Run#", "選用方向", "方向說明", "定稿樣式", "定稿內容", "Tagline", "圖片Prompt"]


def _workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _sheet(wb: Workbook, name: str, headers: list[str], widths: list[int]):
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    ws.append(headers)
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col)].width = w
    return ws


def export_path(cfg) -> Path:
    return Path(cfg.data_dir) / "exports" / "news_ai_records.xlsx"


def existing_run_ids(cfg, sheet_name: str) -> set[int]:
    """讀取某個工作表已匯出的 Run#，避免重複寫入。"""
    path = export_path(cfg)
    if not path.exists():
        return set()
    try:
        wb = load_workbook(path)
    except Exception:  # noqa: BLE001
        return set()
    if sheet_name not in wb.sheetnames:
        return set()
    out: set[int] = set()
    for row in wb[sheet_name].iter_rows(min_row=2, values_only=True):
        v = row[2] if len(row) > 2 else None  # Run# 欄
        if v is not None:
            try:
                out.add(int(v))
            except (TypeError, ValueError):
                pass
    return out


def save_articles(cfg, articles: list[dict], run_id: int, run_date: str, account: str) -> None:
    """把一輪搜到的新聞附加到 Excel「新聞」工作表。"""
    if not getattr(cfg, "excel_export", True):
        return
    path = export_path(cfg)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = _workbook(path)
    ws = _sheet(wb, ARTICLES_SHEET, ARTICLES_HEADERS, [12, 14, 7, 6, 10, 40, 40, 50, 14, 12])
    for a in articles:
        ws.append([
            run_date, account, run_id,
            a.get("id", ""), a.get("topic", ""),
            a.get("title", ""), a.get("url", ""), a.get("snippet", ""),
            a.get("source", ""), a.get("date", ""),
        ])
    wb.save(path)


def save_final(cfg, run_id: int, run_date: str, account: str,
               chosen_direction: str, style: str, script: str,
               tagline: str, image_prompt: str) -> None:
    """把最終定稿（方向 + 腳本 + Tagline + 圖片 Prompt）附加到 Excel「定稿」工作表。"""
    if not getattr(cfg, "excel_export", True):
        return
    try:
        d = json.loads(chosen_direction or "{}") or {}
    except Exception:  # noqa: BLE001
        d = {}
    path = export_path(cfg)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = _workbook(path)
    ws = _sheet(wb, FINAL_SHEET, FINAL_HEADERS, [12, 14, 7, 35, 50, 10, 80, 35, 60])
    ws.append([
        run_date, account, run_id,
        d.get("title", ""), d.get("description", ""),
        style, script, tagline, image_prompt,
    ])
    wb.save(path)
